"""
Excel 自動回填引擎 — 通用 openpyxl 操作

負責將 fill_values 寫入 xlsx 檔案的指定位置，保留字體、對齊、數字格式。
不含任何 domain 邏輯。
"""

import io
import copy
import logging
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, coordinate_to_tuple

from app.autofill_core.field_detection import convert_value

logger = logging.getLogger(__name__)


class ExcelAutoFillEngine:
    """Excel 表單自動回填引擎"""

    async def fill(
        self,
        file_content: bytes,
        field_lookup: dict,
        value_lookup: dict,
    ) -> bytes:
        """將 value_lookup 中的值寫入 field_lookup 指定的位置。

        Args:
            file_content: 原始 xlsx bytes
            field_lookup: {field_id: field_map_entry}
            value_lookup: {field_id: value}

        Returns:
            回填後的 xlsx bytes
        """
        wb = load_workbook(io.BytesIO(file_content))

        for field_id, value in value_lookup.items():
            field = field_lookup.get(field_id)
            if not field:
                continue

            val_loc = field.get("value_location")
            if not val_loc:
                self._fallback_fill_label_cell(wb, field, value)
                continue

            sheet_name = val_loc.get("sheet")
            cell_coord = val_loc.get("cell")

            if not sheet_name or not cell_coord:
                continue
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found, skipping field {field_id}")
                continue

            ws = wb[sheet_name]
            typed_value = convert_value(value, field.get("field_type", "text"))

            target_cell = ws[cell_coord]
            if isinstance(target_cell, MergedCell):
                resolved = self._resolve_merged_cell(ws, cell_coord)
                if resolved:
                    target_cell = ws[resolved]
                    logger.info(f"MergedCell {cell_coord} → 左上角 {resolved}")
                else:
                    logger.warning(f"無法解析合併格 {cell_coord}，跳過")
                    continue

            original_font = copy.copy(target_cell.font) if target_cell.font else None
            original_alignment = copy.copy(target_cell.alignment) if target_cell.alignment else None
            original_number_format = target_cell.number_format

            target_cell.value = typed_value

            if original_font:
                target_cell.font = original_font
            if original_alignment:
                target_cell.alignment = original_alignment
            if original_number_format:
                target_cell.number_format = original_number_format

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _fallback_fill_label_cell(self, wb, field: dict, value) -> None:
        """value_location 不存在時，退回寫入 label 所在儲存格。"""
        label_loc = field.get("label_location", {})
        sheet_name = label_loc.get("sheet")
        cell_coord = label_loc.get("cell")
        if not (sheet_name and cell_coord and sheet_name in wb.sheetnames):
            return

        ws = wb[sheet_name]
        target = ws[cell_coord]
        if isinstance(target, MergedCell):
            resolved = self._resolve_merged_cell(ws, cell_coord)
            if resolved:
                ws[resolved] = value
        else:
            ws[cell_coord] = value

    @staticmethod
    def _resolve_merged_cell(ws, cell_coord: str) -> Optional[str]:
        """找到合併儲存格的左上角座標。"""
        try:
            row, col = coordinate_to_tuple(cell_coord)
        except Exception:
            return None

        for merge_range in ws.merged_cells.ranges:
            if cell_coord in merge_range or (
                merge_range.min_row <= row <= merge_range.max_row
                and merge_range.min_col <= col <= merge_range.max_col
            ):
                return f"{get_column_letter(merge_range.min_col)}{merge_range.min_row}"

        return None
