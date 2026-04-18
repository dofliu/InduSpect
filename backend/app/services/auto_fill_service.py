"""
自動回填服務 — 執行回填、預覽、報告產生

職責：
- 執行 Excel/Word 自動回填（委派給 autofill_core 引擎）
- 預覽回填結果
- 舊版報告產生相容
"""

import io
import logging

import google.generativeai as genai
from openpyxl import load_workbook
from docx import Document

from app.config import settings
from app.autofill_core import ExcelAutoFillEngine, WordAutoFillEngine

logger = logging.getLogger(__name__)


class AutoFillService:
    """自動回填服務 — 薄層包裝，將執行委派給 autofill_core。"""

    def __init__(self):
        genai.configure(api_key=settings.gemini_api_key)
        self._excel_engine = ExcelAutoFillEngine()
        self._word_engine = WordAutoFillEngine()

    # ================================================================
    # 自動回填引擎
    # ================================================================

    async def auto_fill(
        self,
        file_content: bytes,
        file_name: str,
        field_map: list[dict],
        fill_values: list[dict],
    ) -> bytes:
        """執行自動回填：將值寫入原始文件的指定位置。

        Args:
            file_content: 原始文件 bytes
            file_name: 用於判斷格式（xlsx/docx）
            field_map: 欄位位置地圖（含 value_location）
            fill_values: [{"field_id": "...", "value": "..."}, ...]

        Returns:
            回填後的文件 bytes
        """
        file_type = file_name.split('.')[-1].lower()

        value_lookup = {fv["field_id"]: fv["value"] for fv in fill_values}
        field_lookup = {f["field_id"]: f for f in field_map}

        if file_type == 'xlsx':
            return await self._excel_engine.fill(file_content, field_lookup, value_lookup)
        if file_type == 'docx':
            return await self._word_engine.fill(file_content, field_lookup, value_lookup)
        raise ValueError(f"不支援的檔案格式: {file_type}")

    # ================================================================
    # 預覽回填
    # ================================================================

    async def preview_fill(
        self,
        template: dict,
        inspection_data: dict,
    ) -> dict:
        """預覽填入結果（舊版 API 相容）"""
        field_values: dict[str, str] = {}
        warnings: list[str] = []

        for field in template.get("fields", []):
            mapping = field.get("mapping")
            if mapping and mapping in inspection_data:
                value = inspection_data[mapping]
                if isinstance(value, dict):
                    value = str(value)
                field_values[field["field_name"]] = str(value) if value else ""
            else:
                field_values[field["field_name"]] = ""
                if mapping:
                    warnings.append(f"欄位 '{field['field_name']}' 對應的資料不存在")

        return {
            "template_name": template.get("name", ""),
            "vendor_name": template.get("vendor_name", ""),
            "field_values": field_values,
            "warnings": warnings,
        }

    async def preview_auto_fill(
        self,
        field_map: list[dict],
        fill_values: list[dict],
    ) -> dict:
        """預覽自動回填結果（新版 API）— 回傳每欄位預計值與信心度。"""
        value_lookup = {fv["field_id"]: fv for fv in fill_values}

        preview_items: list[dict] = []
        warnings: list[str] = []

        for field in field_map:
            fid = field["field_id"]
            fv = value_lookup.get(fid)

            item = {
                "field_id": fid,
                "field_name": field["field_name"],
                "field_type": field.get("field_type", "text"),
                "value": fv["value"] if fv else None,
                "confidence": fv.get("confidence", 0.0) if fv else 0.0,
                "source": fv.get("source", "") if fv else "",
                "has_target": field.get("value_location") is not None,
            }
            preview_items.append(item)

            if not fv:
                warnings.append(f"欄位 '{field['field_name']}' 無對應值")
            elif item["confidence"] < 0.7:
                warnings.append(
                    f"欄位 '{field['field_name']}' 映射信心度較低 ({item['confidence']:.0%})，建議確認"
                )
            if not item["has_target"]:
                warnings.append(
                    f"欄位 '{field['field_name']}' 找不到值儲存格位置，無法回填"
                )

        return {
            "preview_items": preview_items,
            "total_fields": len(field_map),
            "filled_count": sum(1 for p in preview_items if p["value"] is not None),
            "warnings": warnings,
        }

    # ================================================================
    # 報告生成（舊版 API 相容）
    # ================================================================

    async def generate_report(
        self,
        report_id: str,
        template: dict,
        inspection_data: dict,
        output_format: str = "xlsx",
    ) -> str:
        """產生報告，回傳 output_path"""
        try:
            if template["file_type"] == "xlsx":
                return await self._fill_excel(template, inspection_data, report_id)
            if template["file_type"] == "docx":
                return await self._fill_word(template, inspection_data, report_id)
            raise ValueError(f"Unsupported template type: {template['file_type']}")

        except Exception as e:
            logger.error(f"Generate report failed: {e}")
            raise

    async def _fill_excel(
        self,
        template: dict,
        inspection_data: dict,
        report_id: str,
    ) -> str:
        wb = load_workbook(io.BytesIO(template["file_content"]))
        ws = wb.active

        for field in template["fields"]:
            mapping = field.get("mapping")
            if mapping and mapping in inspection_data:
                ws[field["location"]] = inspection_data[mapping]

        output_path = f"/tmp/report_{report_id}.xlsx"
        wb.save(output_path)
        return output_path

    async def _fill_word(
        self,
        template: dict,
        inspection_data: dict,
        report_id: str,
    ) -> str:
        doc = Document(io.BytesIO(template["file_content"]))

        for para in doc.paragraphs:
            for field in template["fields"]:
                mapping = field.get("mapping")
                if mapping and mapping in inspection_data:
                    value = str(inspection_data[mapping] or "")
                    placeholder = f"{{{{{field['field_name']}}}}}"
                    if placeholder in para.text:
                        para.text = para.text.replace(placeholder, value)

        output_path = f"/tmp/report_{report_id}.docx"
        doc.save(output_path)
        return output_path
